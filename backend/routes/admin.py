# =============================================================
# routes/admin.py — Dashboard privé IKLILOUNE
# Toutes les routes sont protégées par @login_required
# Architecture : 1 section = 1 ressource (Regis N'guessan)
# =============================================================

import os
import json
import io
from datetime import datetime

from flask import (Blueprint, render_template, request,
                   jsonify, redirect, url_for, Response)
from flask_login import login_required, current_user

from backend.database import db
from backend.models.produit      import Produit
from backend.models.commande     import Commande, HistoriqueStatut
from backend.models.client       import Client
from backend.models.banniere     import Banniere
from backend.services.image_service  import traiter_image, supprimer_image
from backend.services.stats_service  import (
    calculer_kpis,
    graphique_ventes_mensuelles,
    graphique_categories,
    graphique_ventes_30_jours,
    ventes_par_mois,
    top_produits_commandes,
    top_produits_consultes,
)
from backend.services.commande_service import message_notification_statut
from backend.models.historique_stock   import HistoriqueStock
from backend import config

admin_bp = Blueprint("admin", __name__)


# ══════════════════════════════════════════════════════════════
# DASHBOARD PRINCIPAL
# ══════════════════════════════════════════════════════════════

@admin_bp.route("/" + config.ADMIN_URL_DASHBOARD)
@login_required
def tableau_de_bord():
    """
    Page principale de l'administration.
    - KPIs en temps réel depuis la base de données
    - Trois graphiques Matplotlib encodés en base64
    - Tops produits (les plus commandés et les plus consultés)
    """
    # KPIs calculés par le service dédié (pas de logique ici)
    kpis = calculer_kpis()

    # Graphiques encodés en base64 PNG (prêts à insérer dans <img src="...">)
    graphique_v  = graphique_ventes_mensuelles()     # barres mensuelles
    graphique_c  = graphique_categories()            # camembert catégories
    graphique_30 = graphique_ventes_30_jours()       # courbe 30 derniers jours

    # Tops produits pour les tableaux du dashboard
    top_commandes  = top_produits_commandes(limite=5)
    top_consultes  = top_produits_consultes(limite=5)

    return render_template(
        "admin/dashboard.html",
        kpis             = kpis,
        graphique_ventes = graphique_v,
        graphique_cats   = graphique_c,
        graphique_30j    = graphique_30,
        top_commandes    = top_commandes,
        top_consultes    = top_consultes,
    )


# ══════════════════════════════════════════════════════════════
# PRODUITS — CRUD complet
# ══════════════════════════════════════════════════════════════

@admin_bp.route("/admin/api/produits")
@login_required
def api_produits_admin():
    """
    Liste TOUS les produits (actifs ET inactifs).
    Utilise vers_dict_admin() pour exposer le stock réel (admin only).
    """
    produits = Produit.query.order_by(Produit.cree_le.desc()).all()
    return jsonify([p.vers_dict_admin() for p in produits])


@admin_bp.route("/admin/api/produit/ajouter", methods=["POST"])
@login_required
def ajouter_produit():
    """
    Ajoute un nouvel article au catalogue.
    Accepte un fichier image qui sera converti en WebP (800×800px max 150Ko).
    La référence IKL-YYYY-XXXXX est générée automatiquement.
    """
    try:
        # ── Validation des champs obligatoires ──────────────────
        nom       = request.form.get("nom", "").strip()
        categorie = request.form.get("categorie", "").strip()
        prix_str  = request.form.get("prix", "0")

        if not nom or not categorie:
            return jsonify({"erreur": "Nom et catégorie obligatoires"}), 400

        try:
            prix = int(prix_str)
            if prix <= 0:
                raise ValueError
        except ValueError:
            return jsonify({"erreur": "Prix invalide (entier positif requis)"}), 400

        prix_promo_val = request.form.get("prix_promo")

        # ── Création du produit ─────────────────────────────────
        produit = Produit(
            nom            = nom,
            categorie      = categorie,
            genre          = request.form.get("genre", "mixte"),
            prix           = prix,
            prix_promo     = int(prix_promo_val) if prix_promo_val else None,
            stock          = int(request.form.get("stock", 0)),
            description    = request.form.get("description", ""),
            badge          = request.form.get("badge", ""),
            en_vedette     = request.form.get("en_vedette", "false").lower() == "true",
            couleurs_json  = request.form.get("couleurs", "[]"),
            tailles_json   = request.form.get("tailles", "[]"),
            attributs_json = request.form.get("attributs", "{}"),
            seuil_bas      = int(request.form.get("seuil_bas", config.SEUIL_STOCK_BAS)),
            seuil_haut     = int(request.form.get("seuil_haut", config.SEUIL_STOCK_HAUT)),
        )
        db.session.add(produit)
        db.session.flush()  # flush pour obtenir l'ID (dans la transaction courante)
        # ── Log stock initial (sécurisé si table absente) ────────
        if produit.stock > 0:
            try:
                db.session.add(HistoriqueStock(
                    produit_id=produit.id,
                    type_mouvement="ajout_initial",
                    quantite_avant=0,
                    quantite_apres=produit.stock,
                    delta=produit.stock,
                    note=f"Création produit : {produit.nom}"
                ))
            except Exception as _e_log:
                print(f"⚠️ Log stock initial ignoré : {_e_log}")
        db.session.commit()  # commit pour obtenir l'ID avant de nommer la photo

        # ── Traitement de la photo (optionnel) ──────────────────
        photo = request.files.get("photo")
        if photo and photo.filename:
            ext = photo.filename.rsplit(".", 1)[-1].lower()
            if ext not in config.FORMATS_ACCEPTES:
                return jsonify({"erreur": f"Format non accepté : {ext}"}), 400

            os.makedirs(config.DOSSIER_TEMP, exist_ok=True)
            chemin_temp = os.path.join(config.DOSSIER_TEMP, f"tmp_{produit.id}.{ext}")
            photo.save(chemin_temp)

            nom_webp = traiter_image(chemin_temp, f"produit_{produit.id}")

            # Nettoyage du fichier temporaire
            if os.path.exists(chemin_temp):
                os.remove(chemin_temp)

            if nom_webp:
                produit.photo = nom_webp
                db.session.commit()

        print(f"✅ Produit ajouté : [{produit.reference}] {produit.nom}")
        return jsonify({"succes": True, "produit": produit.vers_dict_admin()})

    except Exception as e:
        db.session.rollback()
        print(f"❌ Erreur ajouter_produit : {e}")
        return jsonify({"erreur": str(e)}), 500


@admin_bp.route("/admin/api/produit/modifier/<int:pid>", methods=["POST"])
@login_required
def modifier_produit(pid):
    """
    Modifie un article existant.
    Si une nouvelle photo est fournie, l'ancienne est supprimée.
    """
    produit = db.get_or_404(Produit, pid)
    try:
        stock_avant = produit.stock  # mémoriser avant toute modification
        produit.nom            = request.form.get("nom",       produit.nom).strip()
        produit.categorie      = request.form.get("categorie", produit.categorie)
        produit.genre          = request.form.get("genre",     produit.genre)
        produit.prix           = int(request.form.get("prix",  produit.prix))
        produit.stock          = int(request.form.get("stock", produit.stock))
        produit.description    = request.form.get("description", produit.description)
        produit.badge          = request.form.get("badge",     produit.badge)
        produit.couleurs_json  = request.form.get("couleurs",  produit.couleurs_json)
        produit.tailles_json   = request.form.get("tailles",   produit.tailles_json)
        produit.attributs_json = request.form.get("attributs", produit.attributs_json)
        produit.seuil_bas      = int(request.form.get("seuil_bas",  produit.seuil_bas))
        produit.seuil_haut     = int(request.form.get("seuil_haut", produit.seuil_haut))

        en_vedette = request.form.get("en_vedette")
        if en_vedette is not None:
            produit.en_vedette = en_vedette.lower() == "true"

        pp = request.form.get("prix_promo")
        produit.prix_promo = int(pp) if pp else None

        # Remplacement de photo
        photo = request.files.get("photo")
        if photo and photo.filename:
            if produit.photo:
                supprimer_image(produit.photo)  # supprime l'ancienne WebP

            ext = photo.filename.rsplit(".", 1)[-1].lower()
            chemin_temp = os.path.join(config.DOSSIER_TEMP, f"tmp_{produit.id}.{ext}")
            photo.save(chemin_temp)
            nom_webp = traiter_image(chemin_temp, f"produit_{produit.id}")
            if os.path.exists(chemin_temp):
                os.remove(chemin_temp)
            if nom_webp:
                produit.photo = nom_webp

        # ── Log si le stock a été modifié manuellement ───────────
        if produit.stock != stock_avant:
            note_adj = (request.form.get("note_stock", "").strip()
                        or f"Ajustement manuel : {stock_avant} → {produit.stock}")
            try:
                db.session.add(HistoriqueStock(
                    produit_id=produit.id,
                    type_mouvement="ajustement_manuel",
                    quantite_avant=stock_avant,
                    quantite_apres=produit.stock,
                    delta=produit.stock - stock_avant,
                    note=note_adj
                ))
            except Exception as _e_log:
                print(f"⚠️ Log ajustement ignoré : {_e_log}")
        db.session.commit()
        print(f"✅ Produit modifié : [{produit.reference}] {produit.nom}")
        return jsonify({"succes": True, "produit": produit.vers_dict_admin()})

    except Exception as e:
        db.session.rollback()
        print(f"❌ Erreur modifier_produit : {e}")
        return jsonify({"erreur": str(e)}), 500


@admin_bp.route("/admin/api/produit/supprimer/<int:pid>", methods=["DELETE"])
@login_required
def supprimer_produit(pid):
    """
    Désactive un article (soft delete — actif = False).
    Le produit reste en base pour la cohérence des commandes passées.
    """
    produit = db.get_or_404(Produit, pid)
    stock_courant = produit.stock
    produit.actif = False
    try:
        db.session.add(HistoriqueStock(
            produit_id=produit.id,
            type_mouvement="desactivation",
            quantite_avant=stock_courant,
            quantite_apres=stock_courant,
            delta=0,
            note="Produit désactivé — retiré du catalogue"
        ))
    except Exception as _e_log:
        print(f"⚠️ Log désactivation ignoré : {_e_log}")
    db.session.commit()
    print(f"🗑️ Produit désactivé : [{produit.reference}] {produit.nom}")
    return jsonify({"succes": True, "message": f"'{produit.nom}' retiré du catalogue"})


@admin_bp.route("/admin/api/produit/reactiver/<int:pid>", methods=["POST"])
@login_required
def reactiver_produit(pid):
    """Réactive un article précédemment désactivé."""
    produit = db.get_or_404(Produit, pid)
    produit.actif = True
    try:
        db.session.add(HistoriqueStock(
            produit_id=produit.id,
            type_mouvement="reactivation",
            quantite_avant=produit.stock,
            quantite_apres=produit.stock,
            delta=0,
            note="Produit réactivé — remis en ligne"
        ))
    except Exception as _e_log:
        print(f"⚠️ Log réactivation ignoré : {_e_log}")
    db.session.commit()
    return jsonify({"succes": True, "message": f"'{produit.nom}' remis en ligne"})


# ══════════════════════════════════════════════════════════════
# COMMANDES — Gestion et suivi
# ══════════════════════════════════════════════════════════════

@admin_bp.route("/admin/api/commandes")
@login_required
def api_commandes():
    """
    Liste toutes les commandes, de la plus récente à la plus ancienne.
    Inclut l'étape du workflow (0–4) pour les barres de progression.
    """
    commandes = Commande.query.order_by(Commande.cree_le.desc()).all()
    return jsonify([c.vers_dict() for c in commandes])


@admin_bp.route("/admin/api/commande/<int:cid>")
@login_required
def detail_commande(cid):
    """
    Retourne les détails complets d'une commande :
    infos client, articles, totaux, historique des statuts.
    """
    commande = db.get_or_404(Commande, cid)
    data = commande.vers_dict()
    data["historique"] = [h.vers_dict() for h in commande.historique]
    return jsonify({"commande": data})


@admin_bp.route("/admin/api/commande/statut/<int:cid>", methods=["POST"])
@login_required
def modifier_statut_commande(cid):
    """
    Met à jour le statut d'une commande.

    - Enregistre automatiquement un HistoriqueStatut (traçabilité complète)
    - Retourne une URL WhatsApp pré-remplie pour notifier le client
    - Statuts valides : recue → confirmee → en_preparation → expediee → livree | annulee

    Corps JSON attendu :
        {
          "statut": "confirmee",
          "note": "Commande confirmée, paiement reçu"   ← optionnel
        }
    """
    commande = db.get_or_404(Commande, cid)
    data = request.get_json() or {}

    # ── Validation du nouveau statut ────────────────────────────
    statuts_valides = ["recue", "confirmee", "en_preparation",
                       "expediee", "livree", "annulee"]
    nouveau_statut = data.get("statut", "").strip()

    if nouveau_statut not in statuts_valides:
        return jsonify({
            "erreur": f"Statut invalide. Valeurs acceptées : {', '.join(statuts_valides)}"
        }), 400

    statut_avant = commande.statut

    # Pas de changement → rien à faire
    if statut_avant == nouveau_statut:
        return jsonify({"succes": True, "message": "Statut inchangé",
                        "commande": commande.vers_dict()})

    try:
        # ── Mise à jour du statut ────────────────────────────────
        commande.statut = nouveau_statut

        # Note admin optionnelle
        note_admin = data.get("note", "").strip()
        if note_admin:
            commande.notes_admin = note_admin

        # ── Enregistrement de l'historique statut ────────────────
        historique_entry = HistoriqueStatut(
            commande_id  = commande.id,
            statut_avant = statut_avant,
            statut_apres = nouveau_statut,
            note         = note_admin or f"Statut changé de '{statut_avant}' vers '{nouveau_statut}'",
            modifie_par  = current_user.email if current_user.is_authenticated else "système",
        )
        db.session.add(historique_entry)

        # ── Mouvement de stock automatique ────────────────────────
        # CONFIRMATION → décrémenter le stock de chaque article
        if nouveau_statut == "confirmee" and statut_avant != "confirmee":
            try:
                import json as _json
                for article in _json.loads(commande.articles_json or "[]"):
                    pid_art = article.get("id")
                    qty = int(article.get("qty",
                              article.get("quantite",
                              article.get("qte", 1))))
                    if pid_art:
                        prod = db.session.get(Produit, pid_art)
                    else:
                        ref = article.get("reference", "")
                        prod = Produit.query.filter_by(reference=ref).first() if ref else None
                    if prod and qty > 0:
                        avant = prod.stock
                        prod.stock = max(0, prod.stock - qty)
                        db.session.add(HistoriqueStock(
                            produit_id=prod.id,
                            type_mouvement="vente",
                            quantite_avant=avant,
                            quantite_apres=prod.stock,
                            delta=prod.stock - avant,
                            commande_id=commande.id,
                            note=f"Vente — commande {commande.numero}"
                        ))
            except Exception as e_stock:
                print(f"⚠️ Erreur décrément stock : {e_stock}")

        # ANNULATION → remettre le stock uniquement si une vente avait été enregistrée
        elif nouveau_statut == "annulee":
            try:
                vente_log = HistoriqueStock.query.filter_by(
                    commande_id=commande.id,
                    type_mouvement="vente"
                ).first()
            except Exception:
                vente_log = None  # table absente — migration pas encore lancée
            if vente_log:
                try:
                    import json as _json
                    for article in _json.loads(commande.articles_json or "[]"):
                        pid_art = article.get("id")
                        qty = int(article.get("qty",
                                  article.get("quantite",
                                  article.get("qte", 1))))
                        if pid_art:
                            prod = db.session.get(Produit, pid_art)
                        else:
                            ref = article.get("reference", "")
                            prod = Produit.query.filter_by(reference=ref).first() if ref else None
                        if prod and qty > 0:
                            avant = prod.stock
                            prod.stock = prod.stock + qty
                            db.session.add(HistoriqueStock(
                                produit_id=prod.id,
                                type_mouvement="annulation_commande",
                                quantite_avant=avant,
                                quantite_apres=prod.stock,
                                delta=qty,
                                commande_id=commande.id,
                                note=f"Annulation — stock remis ({commande.numero})"
                            ))
                except Exception as e_stock:
                    print(f"⚠️ Erreur récrément stock : {e_stock}")

        db.session.commit()

        # ── Génération du lien WhatsApp pour notifier le client ──
        # message_notification_statut() retourne un dict {"url": ..., "message": ...}
        try:
            notif_wa = message_notification_statut(commande)
        except Exception:
            notif_wa = None  # non bloquant — la mise à jour a bien eu lieu

        print(f"✅ Commande {commande.numero} : {statut_avant} → {nouveau_statut}")

        return jsonify({
            "succes"         : True,
            "commande"       : commande.vers_dict(),
            "notification_wa": notif_wa,   # URL + texte du message client
        })

    except Exception as e:
        db.session.rollback()
        print(f"❌ Erreur modifier_statut_commande : {e}")
        return jsonify({"erreur": str(e)}), 500


@admin_bp.route("/admin/api/commande/notes/<int:cid>", methods=["POST"])
@login_required
def modifier_notes_commande(cid):
    """Met à jour uniquement les notes internes d'une commande (sans changer le statut)."""
    commande = db.get_or_404(Commande, cid)
    data = request.get_json() or {}
    try:
        commande.notes_admin = data.get("notes", commande.notes_admin)
        db.session.commit()
        return jsonify({"succes": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"erreur": str(e)}), 500


# ══════════════════════════════════════════════════════════════
# BANNIÈRES — Carrousel de la page d'accueil
# ══════════════════════════════════════════════════════════════

@admin_bp.route("/admin/api/bannieres")
@login_required
def api_bannieres_admin():
    """
    Liste toutes les bannières (actives et inactives), triées par ordre.
    """
    bannieres = Banniere.query.order_by(Banniere.ordre.asc()).all()
    return jsonify([b.vers_dict() for b in bannieres])


@admin_bp.route("/admin/api/banniere/creer", methods=["POST"])
@login_required
def creer_banniere():
    """
    Crée une nouvelle bannière pour le carrousel d'accueil.

    Corps JSON attendu :
        {
          "titre": "Nouvelle collection Perles",
          "sous_titre": "Découvrez nos parfums d'été",
          "texte_bouton": "Voir la collection",
          "lien_bouton": "/catalogue?genre=femme",
          "collection": "perles",        ← perles | corail | les-deux
          "deco_emoji": "🌸",
          "style": "clair",              ← clair | sombre
          "ordre": 1,
          "date_debut": "2026-06-01",    ← optionnel
          "date_fin": "2026-06-30"       ← optionnel
        }
    """
    try:
        data = request.get_json() or {}

        titre = data.get("titre", "").strip()
        if not titre:
            return jsonify({"erreur": "Le titre est obligatoire"}), 400

        # Parser les dates optionnelles
        def _parse_date(val):
            """Convertit une chaîne YYYY-MM-DD en datetime, ou None si invalide."""
            if not val:
                return None
            try:
                return datetime.strptime(val, "%Y-%m-%d")
            except ValueError:
                return None

        banniere = Banniere(
            titre        = titre,
            sous_titre   = data.get("sous_titre", "").strip(),
            texte_bouton = data.get("texte_bouton", "Découvrir").strip(),
            lien_bouton  = data.get("lien_bouton", "/").strip(),
            collection   = data.get("collection", "les-deux"),
            deco_emoji   = data.get("deco_emoji", "✨"),
            style        = data.get("style", "clair"),
            ordre        = int(data.get("ordre", 99)),
            date_debut   = _parse_date(data.get("date_debut")),
            date_fin     = _parse_date(data.get("date_fin")),
            actif        = True,
        )
        db.session.add(banniere)
        db.session.commit()

        print(f"✅ Bannière créée : [{banniere.id}] {banniere.titre}")
        return jsonify({"succes": True, "banniere": banniere.vers_dict()}), 201

    except Exception as e:
        db.session.rollback()
        print(f"❌ Erreur creer_banniere : {e}")
        return jsonify({"erreur": str(e)}), 500


@admin_bp.route("/admin/api/banniere/modifier/<int:bid>", methods=["POST"])
@login_required
def modifier_banniere(bid):
    """
    Modifie une bannière existante.
    Tous les champs sont optionnels — seuls les champs fournis sont mis à jour.
    """
    banniere = db.get_or_404(Banniere, bid)
    try:
        data = request.get_json() or {}

        def _parse_date(val):
            if not val:
                return None
            try:
                return datetime.strptime(val, "%Y-%m-%d")
            except ValueError:
                return None

        # Mise à jour champ par champ (valeur actuelle si absent du JSON)
        if "titre"        in data: banniere.titre        = data["titre"].strip()
        if "sous_titre"   in data: banniere.sous_titre   = data["sous_titre"].strip()
        if "texte_bouton" in data: banniere.texte_bouton = data["texte_bouton"].strip()
        if "lien_bouton"  in data: banniere.lien_bouton  = data["lien_bouton"].strip()
        if "collection"   in data: banniere.collection   = data["collection"]
        if "deco_emoji"   in data: banniere.deco_emoji   = data["deco_emoji"]
        if "style"        in data: banniere.style        = data["style"]
        if "ordre"        in data: banniere.ordre        = int(data["ordre"])
        if "actif"        in data: banniere.actif        = bool(data["actif"])
        if "date_debut"   in data: banniere.date_debut   = _parse_date(data["date_debut"])
        if "date_fin"     in data: banniere.date_fin     = _parse_date(data["date_fin"])

        db.session.commit()
        return jsonify({"succes": True, "banniere": banniere.vers_dict()})

    except Exception as e:
        db.session.rollback()
        return jsonify({"erreur": str(e)}), 500


@admin_bp.route("/admin/api/banniere/reordonner", methods=["POST"])
@login_required
def reordonner_bannieres():
    """
    Met à jour l'ordre d'affichage de plusieurs bannières en une seule requête.

    Corps JSON attendu :
        {"ordre": [{"id": 3, "ordre": 1}, {"id": 1, "ordre": 2}, ...]}
    """
    try:
        data = request.get_json() or {}
        items = data.get("ordre", [])

        for item in items:
            bid   = item.get("id")
            ordre = item.get("ordre")
            if bid is not None and ordre is not None:
                banniere = db.session.get(Banniere, bid)
                if banniere:
                    banniere.ordre = int(ordre)

        db.session.commit()
        return jsonify({"succes": True, "message": f"{len(items)} bannières réordonnées"})

    except Exception as e:
        db.session.rollback()
        return jsonify({"erreur": str(e)}), 500


@admin_bp.route("/admin/api/banniere/supprimer/<int:bid>", methods=["DELETE"])
@login_required
def supprimer_banniere(bid):
    """
    Désactive une bannière (soft delete).
    La bannière reste en base pour l'historique des campagnes.
    """
    banniere = db.get_or_404(Banniere, bid)
    banniere.actif = False
    db.session.commit()
    return jsonify({"succes": True, "message": f"Bannière '{banniere.titre}' désactivée"})


# ══════════════════════════════════════════════════════════════
# CLIENTS — Gestion du registre
# ══════════════════════════════════════════════════════════════

@admin_bp.route("/admin/api/clients")
@login_required
def api_clients():
    """Liste tous les clients, du plus récent au plus ancien."""
    clients = Client.query.order_by(Client.cree_le.desc()).all()
    return jsonify([c.vers_dict() for c in clients])


@admin_bp.route("/admin/api/clients/export-csv")
@login_required
def exporter_clients_csv():
    """
    Exporte le registre clients en CSV (encodage UTF-8 avec BOM pour Excel).
    Colonnes : id, prénom, nom, téléphone, email, nb_commandes, source, date_inscription.
    """
    import csv

    clients = Client.query.filter_by(actif=True).order_by(Client.cree_le.desc()).all()
    output  = io.StringIO()

    # Définir les colonnes exportées
    colonnes = ["id", "prenom", "nom", "telephone", "email",
                "nb_commandes", "interet", "source", "cree_le"]
    writer = csv.DictWriter(output, fieldnames=colonnes, extrasaction="ignore")
    writer.writeheader()

    for c in clients:
        writer.writerow(c.vers_dict())

    # BOM UTF-8 pour que Excel ouvre le CSV correctement
    return Response(
        "﻿" + output.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={
            "Content-Disposition":
            f"attachment; filename=clients_ikliloune_{datetime.now().strftime('%Y%m%d')}.csv"
        }
    )


@admin_bp.route("/admin/api/client/ajouter", methods=["POST"])
@login_required
def ajouter_client_manuel():
    """
    Ajoute un client manuellement depuis l'admin.
    Utilisé pour les clients reçus par téléphone, WhatsApp ou en boutique.
    Le téléphone est l'identifiant principal — l'email est optionnel.
    """
    try:
        data = request.get_json() or {}

        telephone = data.get("telephone", "").strip()
        if not telephone:
            return jsonify({"erreur": "Numéro de téléphone obligatoire"}), 400

        # Vérification unicité téléphone
        if Client.query.filter_by(telephone=telephone).first():
            return jsonify({"erreur": "Ce numéro de téléphone est déjà enregistré"}), 400

        client = Client(
            prenom         = data.get("prenom", "").strip(),
            nom            = data.get("nom", "").strip(),
            email          = data.get("email", "").strip() or None,
            telephone      = telephone,
            interet        = data.get("interet", "tout"),
            adresse        = data.get("adresse", "").strip(),
            source         = "manuel",
            nb_commandes   = int(data.get("nb_commandes", 0)),
            consentement   = bool(data.get("consentement", False)),
            actif          = True,
        )
        db.session.add(client)
        db.session.commit()

        print(f"✅ Client ajouté : {client.prenom} {client.nom} ({client.telephone})")
        return jsonify({"succes": True, "client": client.vers_dict()}), 201

    except Exception as e:
        db.session.rollback()
        print(f"❌ Erreur ajouter_client_manuel : {e}")
        return jsonify({"erreur": str(e)}), 500


@admin_bp.route("/admin/api/client/modifier/<int:cid>", methods=["POST"])
@login_required
def modifier_client(cid):
    """Met à jour les informations d'un client existant."""
    client = db.get_or_404(Client, cid)
    try:
        data = request.get_json() or {}
        if "prenom"         in data: client.prenom         = data["prenom"].strip()
        if "nom"            in data: client.nom            = data["nom"].strip()
        if "telephone"      in data: client.telephone      = data["telephone"].strip()
        if "email"          in data: client.email          = data["email"].strip() or None
        if "interet"        in data: client.interet        = data["interet"]
        if "adresse"        in data: client.adresse        = data["adresse"].strip()
        if "nb_commandes"   in data: client.nb_commandes   = int(data["nb_commandes"])
        if "actif"          in data: client.actif          = bool(data["actif"])
        if "consentement"   in data: client.consentement   = bool(data["consentement"])

        db.session.commit()
        return jsonify({"succes": True, "client": client.vers_dict()})

    except Exception as e:
        db.session.rollback()
        return jsonify({"erreur": str(e)}), 500


# ══════════════════════════════════════════════════════════════
# CODES PROMO — Gestion des campagnes de réduction
# ══════════════════════════════════════════════════════════════

@admin_bp.route("/admin/api/codes-promo")
@login_required
def api_codes_promo():
    """Liste tous les codes promo, du plus récent au plus ancien."""
    from backend.models.code_promo import CodePromo
    codes = CodePromo.query.order_by(CodePromo.cree_le.desc()).all()
    return jsonify([c.vers_dict() for c in codes])


@admin_bp.route("/admin/api/code-promo/creer", methods=["POST"])
@login_required
def creer_code_promo():
    """
    Crée un nouveau code promo.

    Corps JSON :
        {
          "code": "ETE2026",
          "description": "Soldes d'été",
          "type_reduction": "pourcentage",   ← pourcentage | montant_fixe
          "reduction_pct": 15,               ← si pourcentage
          "montant_min": 5000,               ← panier minimum en FCFA (optionnel)
          "max_utilisations": 100,           ← null = illimité
          "conditions": "tous",              ← tous | nouveaux_clients | evenement
          "date_debut": "2026-06-01",        ← optionnel
          "expire_le": "2026-08-31"          ← optionnel
        }
    """
    from backend.models.code_promo import CodePromo

    try:
        data = request.get_json() or {}
        code_str = data.get("code", "").strip().upper()

        if not code_str:
            return jsonify({"erreur": "Le code est obligatoire"}), 400

        if CodePromo.query.filter_by(code=code_str).first():
            return jsonify({"erreur": f"Le code '{code_str}' existe déjà"}), 400

        # Conversion des dates
        def _parse_date(val):
            if not val:
                return None
            try:
                return datetime.strptime(val, "%Y-%m-%d")
            except ValueError:
                return None

        code = CodePromo(
            code             = code_str,
            description      = data.get("description", ""),
            type_reduction   = data.get("type_reduction", "pourcentage"),
            reduction_pct    = int(data.get("reduction_pct", 5)),
            montant_min      = int(data.get("montant_min", 0)) or None,
            max_utilisations = int(data["max_utilisations"]) if data.get("max_utilisations") else None,
            conditions       = data.get("conditions", "tous"),
            date_debut       = _parse_date(data.get("date_debut")),
            expire_le        = _parse_date(data.get("expire_le")),
            actif            = True,
        )
        db.session.add(code)
        db.session.commit()

        print(f"✅ Code promo créé : {code.code} (-{code.reduction_pct}%)")
        return jsonify({"succes": True, "code": code.vers_dict()}), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({"erreur": str(e)}), 500


@admin_bp.route("/admin/api/code-promo/desactiver/<int:cid>", methods=["POST"])
@login_required
def desactiver_code_promo(cid):
    """Désactive un code promo (soft delete)."""
    from backend.models.code_promo import CodePromo
    code = db.get_or_404(CodePromo, cid)
    code.actif = False
    db.session.commit()
    return jsonify({"succes": True, "message": f"Code '{code.code}' désactivé"})


# ══════════════════════════════════════════════════════════════
# STATS ET EXPORT — Rapports et tableaux Excel
# ══════════════════════════════════════════════════════════════

@admin_bp.route("/admin/api/stats/ventes")
@login_required
def api_stats_ventes():
    """
    Retourne les ventes par mois pour une année donnée.
    Paramètre URL : ?annee=2026 (défaut = année courante)
    """
    annee = request.args.get("annee", type=int, default=datetime.now().year)
    return jsonify({
        "annee"  : annee,
        "mois"   : ventes_par_mois(annee),
    })


@admin_bp.route("/admin/api/stats/kpis")
@login_required
def api_stats_kpis():
    """Retourne les KPIs en temps réel (JSON) — pour les mises à jour AJAX du dashboard."""
    return jsonify(calculer_kpis())


@admin_bp.route("/admin/api/export/ventes-excel")
@login_required
def exporter_ventes_excel():
    """
    Exporte les ventes en fichier Excel (.xlsx) via openpyxl.
    Feuilles :
        1. Résumé mensuel (CA par mois pour l'année sélectionnée)
        2. Toutes les commandes (avec client, articles, total, statut)
        3. Top produits (par nombre de commandes)

    Paramètre URL : ?annee=2026 (défaut = année courante)
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({
            "erreur": "openpyxl non installé. Exécutez : pip install openpyxl"
        }), 500

    annee = request.args.get("annee", type=int, default=datetime.now().year)

    wb = openpyxl.Workbook()

    # ── Style commun ────────────────────────────────────────────
    ROSE_HEADER = "FFE8EEF5"   # bleu-gris clair (ton IKLILOUNE)
    OR_HEADER   = "FFF0C96B"   # or IKLILOUNE
    FONT_TITRE  = Font(name="Calibri", bold=True, size=14, color="FF1A1A2E")
    FONT_EN_T   = Font(name="Calibri", bold=True, size=10, color="FF1A1A2E")
    FONT_NORMAL = Font(name="Calibri", size=10)

    def _style_entete(ws, row=1):
        """Applique un style doré aux cellules de l'en-tête."""
        for cell in ws[row]:
            cell.font      = FONT_EN_T
            cell.fill      = PatternFill("solid", fgColor=OR_HEADER)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_largeur(ws):
        """Ajuste automatiquement la largeur des colonnes."""
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    # ─────────────────────────────────────────────────────────────
    # Feuille 1 : Résumé mensuel CA
    # ─────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = f"CA Mensuel {annee}"

    ws1.append([f"Chiffre d'Affaires Mensuel — IKLILOUNE {annee}"])
    ws1["A1"].font = FONT_TITRE
    ws1.merge_cells("A1:C1")
    ws1.append([])  # ligne vide

    ws1.append(["Mois", "CA (FCFA)", "Nb commandes"])
    _style_entete(ws1, row=3)

    mois_labels = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                   "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]

    donnees_mois = ventes_par_mois(annee)  # liste de 12 valeurs CA

    # Récupérer le nombre de commandes par mois pour l'année
    from sqlalchemy import extract, func as sqlfunc
    nb_par_mois = (
        db.session.query(
            extract("month", Commande.cree_le).label("mois"),
            sqlfunc.count(Commande.id).label("nb")
        )
        .filter(extract("year", Commande.cree_le) == annee)
        .filter(Commande.statut.notin_(["annulee"]))
        .group_by("mois")
        .all()
    )
    nb_dict = {int(row.mois): int(row.nb) for row in nb_par_mois}

    total_ca = 0
    for i, ca in enumerate(donnees_mois, start=1):
        nb = nb_dict.get(i, 0)
        ws1.append([mois_labels[i - 1], ca, nb])
        total_ca += ca

    # Ligne total
    ws1.append(["TOTAL ANNUEL", total_ca, sum(nb_dict.values())])
    derniere = ws1.max_row
    for cell in ws1[derniere]:
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.fill = PatternFill("solid", fgColor=ROSE_HEADER)

    _auto_largeur(ws1)

    # ─────────────────────────────────────────────────────────────
    # Feuille 2 : Détail des commandes
    # ─────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Commandes")

    ws2.append(["Numéro", "Date", "Client", "Téléphone",
                "Articles", "Sous-total", "Remise", "Total",
                "Statut", "Canal", "Notes"])
    _style_entete(ws2)

    commandes_annee = (
        Commande.query
        .filter(extract("year", Commande.cree_le) == annee)
        .order_by(Commande.cree_le.desc())
        .all()
    )

    for c in commandes_annee:
        # Résumé des articles : "Parfum Oud x2, Sac Python x1"
        try:
            articles = json.loads(c.articles_json or "[]")
            resume_art = ", ".join(
                f"{a.get('nom','?')} x{a.get('quantite',1)}" for a in articles
            )
        except Exception:
            resume_art = c.articles_json or ""

        ws2.append([
            c.numero,
            c.cree_le.strftime("%d/%m/%Y %H:%M") if c.cree_le else "",
            f"{c.client_prenom or ''} {c.client_nom}".strip(),
            c.client_telephone,
            resume_art,
            c.sous_total or c.total,
            c.remise_montant or 0,
            c.total,
            c.libelle_statut(),
            c.canal or "site",
            c.notes_admin or "",
        ])

    _auto_largeur(ws2)

    # ─────────────────────────────────────────────────────────────
    # Feuille 3 : Top produits commandés
    # ─────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Top Produits")

    ws3.append(["Rang", "Référence", "Produit", "Catégorie",
                "Genre", "Prix", "Nb commandes", "Nb consultations", "Stock"])
    _style_entete(ws3)

    top = top_produits_commandes(limite=50)  # liste de dicts
    for rang, p in enumerate(top, start=1):
        ws3.append([
            rang,
            p.get("reference", ""),
            p.get("nom", ""),
            p.get("categorie", ""),
            p.get("genre", ""),
            p.get("prix", 0),
            p.get("nb_commandes", 0),
            p.get("nb_consultations", 0),
            p.get("stock", ""),  # stock réel visible uniquement depuis admin
        ])

    _auto_largeur(ws3)

    # ── Sauvegarder et retourner le fichier ─────────────────────
    output_bytes = io.BytesIO()
    wb.save(output_bytes)
    output_bytes.seek(0)

    nom_fichier = f"ikliloune_ventes_{annee}.xlsx"
    return Response(
        output_bytes.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nom_fichier}"}
    )


@admin_bp.route("/admin/api/export/commandes-excel")
@login_required
def exporter_commandes_excel():
    """
    Export Excel de toutes les commandes (toutes années).
    Filtre optionnel : ?statut=livree
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"erreur": "openpyxl non installé"}), 500

    statut_filtre = request.args.get("statut", "")

    requete = Commande.query.order_by(Commande.cree_le.desc())
    if statut_filtre:
        requete = requete.filter(Commande.statut == statut_filtre)

    commandes = requete.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commandes IKLILOUNE"

    OR_HEADER = "FFF0C96B"
    ws.append(["Numéro", "Date", "Prénom", "Nom", "Téléphone", "Adresse",
               "Articles", "Sous-total", "Remise", "Code promo",
               "Total FCFA", "Statut", "Canal", "Notes"])

    for cell in ws[1]:
        cell.font      = Font(bold=True, size=10)
        cell.fill      = PatternFill("solid", fgColor=OR_HEADER)
        cell.alignment = Alignment(horizontal="center")

    for c in commandes:
        try:
            articles = json.loads(c.articles_json or "[]")
            resume   = ", ".join(f"{a.get('nom','?')} x{a.get('quantite',1)}" for a in articles)
        except Exception:
            resume = c.articles_json or ""

        ws.append([
            c.numero,
            c.cree_le.strftime("%d/%m/%Y") if c.cree_le else "",
            c.client_prenom or "",
            c.client_nom,
            c.client_telephone,
            c.client_adresse or "",
            resume,
            c.sous_total or c.total,
            c.remise_montant or 0,
            c.code_promo_utilise or "",
            c.total,
            c.libelle_statut(),
            c.canal or "site",
            c.notes_admin or "",
        ])

    # Largeur automatique
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    output_bytes = io.BytesIO()
    wb.save(output_bytes)
    output_bytes.seek(0)

    suffix = f"_{statut_filtre}" if statut_filtre else ""
    nom = f"commandes_ikliloune{suffix}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return Response(
        output_bytes.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nom}"}
    )


# ══════════════════════════════════════════════════════════════
# LEADS — Pop-up de capture email (-5%)
# Route publique (pas de @login_required)
# ══════════════════════════════════════════════════════════════

@admin_bp.route("/api/lead", methods=["POST"])
def enregistrer_lead():
    """
    Enregistre un prospect depuis le pop-up de bienvenue (-5%).
    Route PUBLIQUE — accessible sans connexion admin.
    Le téléphone est facultatif ici (capture rapide).
    """
    try:
        data  = request.get_json() or {}
        email = data.get("email", "").strip()

        if not email:
            return jsonify({"erreur": "Email obligatoire pour obtenir le code"}), 400

        # Si déjà inscrit, on renvoie simplement son code
        existant = Client.query.filter_by(email=email).first()
        if existant:
            return jsonify({
                "succes"  : True,
                "code"    : "IKLI5",
                "message" : "Vous êtes déjà inscrit(e) ! Votre code -5% : IKLI5"
            })

        # Enregistrement du nouveau lead
        client = Client(
            prenom       = data.get("prenom", "").strip(),
            nom          = data.get("nom", "").strip(),
            email        = email,
            telephone    = data.get("telephone", "").strip() or None,
            interet      = data.get("interet", ""),
            source       = "popup",
            consentement = True,   # a cliqué sur "recevoir le code" = consentement
        )
        db.session.add(client)
        db.session.commit()

        print(f"📧 Nouveau lead : {client.prenom} {client.email}")
        return jsonify({"succes": True, "code": "IKLI5"})

    except Exception as e:
        db.session.rollback()
        return jsonify({"erreur": str(e)}), 500


# ══════════════════════════════════════════════════════════════
# HISTORIQUE STOCK — Audit et traçabilité sécurisée
# ══════════════════════════════════════════════════════════════

@admin_bp.route("/admin/api/historique-stock")
@login_required
def api_historique_stock():
    """
    Tous les mouvements de stock, du plus récent au plus ancien.
    Filtres optionnels : ?type=vente|ajustement_manuel|... & ?limite=200
    """
    type_filtre = request.args.get("type", "")
    limite      = request.args.get("limite", 200, type=int)

    q = HistoriqueStock.query.order_by(HistoriqueStock.date_mouvement.desc())
    if type_filtre:
        q = q.filter(HistoriqueStock.type_mouvement == type_filtre)

    return jsonify([m.vers_dict() for m in q.limit(limite).all()])


@admin_bp.route("/admin/api/historique-stock/<int:pid>")
@login_required
def api_historique_stock_produit(pid):
    """Tous les mouvements pour un produit donné + infos du produit."""
    produit    = db.get_or_404(Produit, pid)
    mouvements = (
        HistoriqueStock.query
        .filter_by(produit_id=pid)
        .order_by(HistoriqueStock.date_mouvement.desc())
        .all()
    )
    return jsonify({
        "produit"    : produit.vers_dict_admin(),
        "mouvements" : [m.vers_dict() for m in mouvements],
    })


# ══════════════════════════════════════════════════════════════
# CAISSE MAGASIN — Ventes en boutique physique
# ══════════════════════════════════════════════════════════════

@admin_bp.route("/admin/api/commande-magasin", methods=["POST"])
@login_required
def creer_commande_magasin():
    """
    Crée une vente en boutique physique (canal = "magasin").
    La commande est directement confirmée et le stock est décrémenté.

    Corps JSON :
    {
      "client_nom":    "Fatou Koné",
      "client_tel":    "+2250700000000",
      "articles":      [{"produit_id": 5, "quantite": 2}, ...],
      "mode_paiement": "orange_money"   ← optionnel
    }
    """
    from backend.models.historique_stock import HistoriqueStock

    data = request.get_json() or {}

    client_nom = data.get("client_nom", "").strip()
    client_tel = data.get("client_tel", "").strip()
    articles_in = data.get("articles", [])

    if not client_nom or not client_tel:
        return jsonify({"erreur": "Nom et téléphone client obligatoires"}), 400
    if not articles_in:
        return jsonify({"erreur": "Aucun article sélectionné"}), 400

    try:
        import json as _json

        # ── Constituer le panier ─────────────────────────────
        articles_json_list = []
        sous_total = 0

        for item in articles_in:
            pid_  = item.get("produit_id")
            qty_  = int(item.get("quantite", item.get("qty", 1)))
            if not pid_ or qty_ <= 0:
                continue
            prod_ = db.session.get(Produit, pid_)
            if not prod_ or not prod_.actif:
                return jsonify({"erreur": f"Produit ID {pid_} introuvable ou inactif"}), 400
            if prod_.stock < qty_:
                return jsonify({
                    "erreur": f"Stock insuffisant pour '{prod_.nom}' "
                              f"(disponible : {prod_.stock}, demandé : {qty_})"
                }), 400

            prix_u = prod_.prix_actuel()
            articles_json_list.append({
                "id"          : prod_.id,
                "reference"   : prod_.reference,
                "nom"         : prod_.nom,
                "prix_actuel" : prix_u,
                "prix_unitaire": prix_u,
                "quantite"    : qty_,
                "qty"         : qty_,
                "photo"       : prod_.photo or "",
                "categorie"   : prod_.categorie,
            })
            sous_total += prix_u * qty_

        if not articles_json_list:
            return jsonify({"erreur": "Aucun article valide"}), 400

        # ── Créer la commande magasin ─────────────────────────
        commande = Commande(
            client_nom       = client_nom,
            client_telephone = client_tel,
            articles_json    = _json.dumps(articles_json_list, ensure_ascii=False),
            sous_total       = sous_total,
            remise_montant   = 0,
            total            = sous_total,
            canal            = "magasin",
            mode_livraison   = "click_collect",
            frais_livraison  = 0,
            statut           = "confirmee",
            mode_paiement    = data.get("mode_paiement", "a_definir"),
            notes_admin      = f"Vente en boutique — saisie par {current_user.email}",
        )
        db.session.add(commande)
        db.session.flush()  # obtenir l'ID avant le log

        # ── Décrémenter le stock ──────────────────────────────
        for item in articles_json_list:
            prod_ = db.session.get(Produit, item["id"])
            if prod_:
                avant_ = prod_.stock
                prod_.stock = max(0, prod_.stock - item["qty"])
                db.session.add(HistoriqueStock(
                    produit_id=prod_.id,
                    type_mouvement="vente_magasin",
                    quantite_avant=avant_,
                    quantite_apres=prod_.stock,
                    delta=prod_.stock - avant_,
                    commande_id=commande.id,
                    note=f"Vente boutique — {commande.numero}"
                ))

        # ── Historique statut ─────────────────────────────────
        db.session.add(HistoriqueStatut(
            commande_id  = commande.id,
            statut_avant = None,
            statut_apres = "confirmee",
            note         = "Vente créée et confirmée en boutique",
            modifie_par  = current_user.email,
        ))

        db.session.commit()

        # ── Générer le ticket WhatsApp acheteur ───────────────
        from backend.services.commande_service import generer_ticket_acheteur
        ticket = generer_ticket_acheteur(commande)

        print(f"🛍️ Vente magasin : {commande.numero} — {client_nom} — {sous_total} FCFA")
        return jsonify({
            "succes"    : True,
            "commande"  : commande.vers_dict(),
            "ticket_wa" : ticket,
        }), 201

    except Exception as e:
        db.session.rollback()
        print(f"❌ Erreur creer_commande_magasin : {e}")
        return jsonify({"erreur": str(e)}), 500


@admin_bp.route("/admin/api/ticket-wa/<int:cid>")
@login_required
def api_ticket_whatsapp_acheteur(cid):
    """
    Génère le ticket WhatsApp acheteur pour n'importe quelle commande.
    Utilisé par le bouton 'Envoyer ticket acheteur' dans le modal commande.
    """
    from backend.services.commande_service import generer_ticket_acheteur
    commande = db.get_or_404(Commande, cid)
    ticket   = generer_ticket_acheteur(commande)
    return jsonify({"succes": True, **ticket})
